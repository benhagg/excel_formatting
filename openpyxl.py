#Group Members: Ryan Briggs, Ben Haggard, Ian Jones, Josh Gillespie
'''We are assuming a high school teacher approached our group complaining about the excel file their grading system produces each quarter.
It spits out all the classes they teach with all their students in a single spreadsheet,
with student information stored in a single column.
# The teacher wants our group to make a program that will automatically format and summarize the important information
about each of the classes they teach.'''
# importing excel functions

import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook

# importing excel file with relative path
external_workbook = openpyxl.load_workbook(r"Poorly_Organized_Data_1(1).xlsx")

# Creating new work book
edited_workbook = openpyxl.Workbook()

# Iterating through each sheet in loaded workbook
for sheet_name in external_workbook.sheetnames:

# selecting current sheet in loaded workbook
sheet = external_workbook[sheet_name]

# this tracks the unique class names
classes = set()

# iterates through each row to find unique class names
for row in sheet.iter_rows(min_row = 2, max_col= 1, max_row= sheet.max_row, values_only= True):

    classes.add(row[0])

# copy data for each class to the correct sheet
for class_name in classes:

    class_sheet = edited_workbook.create_sheet(title = class_name)

    class_sheet.append(["First Name", "Last Name", "Student ID", "Grade"])

    # goes through each row to find data for each class
    for row in sheet.iter_rows(min_row= 2, values_only= True):

        if row[0] == class_name:
            # Split student info into separate parts
            student_info = row[1].split("_")
            last_name = student_info[0]
            first_name = student_info[1]
            student_id = student_info[2]
            grade = row[2]
            class_sheet.append([last_name, first_name, student_id, grade])
# removes unwanted sheet     
edited_workbook.remove(edited_workbook["Sheet"])
saves workbook
edited_workbook.save("Edited_Workbook.xlsx")
